"""
app.py — Cruce CDR vs Proveedores
Streamlit app — despliega gratis en streamlit.io/cloud
"""

import re
import io
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ── Colores ───────────────────────────────────────────────────────────────────
FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")
FILL_BLUE   = PatternFill("solid", fgColor="BDD7EE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FONT_RED    = Font(color="9C0006", bold=True)
FONT_BLUE   = Font(color="1F4E79", bold=True)
FONT_YELLOW = Font(color="9C6500", bold=True)
FILL_HDR    = PatternFill("solid", fgColor="4472C4")
FONT_HDR    = Font(color="FFFFFF", bold=True)

PRICE_KW = ["precio","price","pvp","costo","valor","neto","unitario","venta"]
BAR_KW   = ["barra","barcode","ean","cod","codigo","gtin","upc"]
DESC_KW  = ["desc","nombre","name","product","articulo","item","detalle"]

# ── Helpers ───────────────────────────────────────────────────────────────────
def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s).lower().strip())

def find_cols(headers, keywords):
    return [i for i, h in enumerate(headers)
            if any(k in norm(str(h)) for k in keywords)]

def to_num(v):
    try:
        return float(re.sub(r"[^\d.-]", "", str(v)))
    except Exception:
        return None

def fmt_money(v):
    n = to_num(v)
    if n is None:
        return str(v) if v not in (None, "") else ""
    return f"${n:,.0f}".replace(",", ".")

def read_df(file_bytes):
    bio = io.BytesIO(file_bytes)
    try:
        df = pd.read_excel(bio, dtype=str, engine="openpyxl")
    except Exception:
        bio.seek(0)
        df = pd.read_excel(bio, dtype=str, engine="xlrd")
    df.fillna("", inplace=True)
    return df

def find_match(cdr_row, cdr_bar_cols, cdr_desc_cols,
               prov_df, prov_bar_cols, prov_desc_cols):
    cdr_bars  = {norm(cdr_row.iloc[i]) for i in cdr_bar_cols  if norm(cdr_row.iloc[i])}
    cdr_descs = {norm(cdr_row.iloc[i]) for i in cdr_desc_cols if norm(cdr_row.iloc[i])}

    if cdr_bars and prov_bar_cols:
        for pi, prow in prov_df.iterrows():
            prov_bars = {norm(prow.iloc[i]) for i in prov_bar_cols if norm(prow.iloc[i])}
            if cdr_bars & prov_bars:
                return pi

    if cdr_descs and prov_desc_cols:
        for pi, prow in prov_df.iterrows():
            prov_descs = {norm(prow.iloc[i]) for i in prov_desc_cols if norm(prow.iloc[i])}
            for cd in cdr_descs:
                for pd_ in prov_descs:
                    if len(cd) >= 4 and len(pd_) >= 4:
                        if cd in pd_ or pd_ in cd:
                            return pi
    return None

def compare_prices(cdr_row, cdr_price_cols, prov_row, prov_price_cols):
    cp = [n for i in cdr_price_cols  if (n := to_num(cdr_row.iloc[i]))  is not None]
    pp = [n for i in prov_price_cols if (n := to_num(prov_row.iloc[i])) is not None]
    if not cp or not pp:
        return "sin_precio", None, None
    ca, pa = sum(cp) / len(cp), sum(pp) / len(pp)
    if ca > pa + 0.01:   return "mas_caro",   ca, pa
    elif ca < pa - 0.01: return "mas_barato", ca, pa
    else:                return "igual",       ca, pa

# ── Proceso principal ─────────────────────────────────────────────────────────
def procesar(uploaded_files):
    cdr_file   = next((f for f in uploaded_files if "CDR" in f.name.upper()), None)
    prov_files = [f for f in uploaded_files if "CDR" not in f.name.upper()]

    if not cdr_file:
        st.error("No se encontró archivo con 'CDR' en el nombre.")
        return None
    if not prov_files:
        st.error("No se encontraron archivos de proveedores.")
        return None

    st.info(f"**Base CDR:** {cdr_file.name}")
    for p in prov_files:
        st.info(f"**Proveedor:** {p.name}")

    # Leer CDR
    cdr_df         = read_df(cdr_file.read())
    cdr_headers    = list(cdr_df.columns)
    cdr_bar_cols   = find_cols(cdr_headers, BAR_KW)
    cdr_desc_cols  = find_cols(cdr_headers, DESC_KW)
    cdr_price_cols = find_cols(cdr_headers, PRICE_KW)
    st.write(f"CDR: **{len(cdr_df)} filas** | "
             f"Columnas barras: `{[cdr_headers[i] for i in cdr_bar_cols]}` | "
             f"Columnas precios: `{[cdr_headers[i] for i in cdr_price_cols]}`")

    # Leer proveedores
    proveedores = []
    for pf in prov_files:
        df   = read_df(pf.read())
        hdrs = list(df.columns)
        proveedores.append({
            "name":       re.sub(r"\.[^.]+$", "", pf.name),
            "df":         df,
            "headers":    hdrs,
            "bar_cols":   find_cols(hdrs, BAR_KW),
            "desc_cols":  find_cols(hdrs, DESC_KW),
            "price_cols": find_cols(hdrs, PRICE_KW),
        })
        st.write(f"{pf.name}: **{len(df)} filas** | "
                 f"Barras: `{[hdrs[i] for i in find_cols(hdrs, BAR_KW)]}` | "
                 f"Precios: `{[hdrs[i] for i in find_cols(hdrs, PRICE_KW)]}`")

    # Cruce
    progress = st.progress(0, text="Cruzando datos...")
    out_headers = [f"CDR: {h}" for h in cdr_headers]
    for pv in proveedores:
        out_headers += [f"{pv['name']}: {h}" for h in pv["headers"]]
    out_headers += ["Estado Precio", "Precio CDR", "Precio Proveedor"]

    out_rows, row_meta = [], []
    total = len(cdr_df)

    for idx, (_, cdr_row) in enumerate(cdr_df.iterrows()):
        result_row        = list(cdr_row)
        matched           = False
        price_status      = "sin_match"
        cdr_p = prov_p   = None
        price_col_indices = set(cdr_price_cols)
        offset            = len(cdr_headers)

        for pv in proveedores:
            mi = find_match(cdr_row, cdr_bar_cols, cdr_desc_cols,
                            pv["df"], pv["bar_cols"], pv["desc_cols"])
            if mi is not None:
                prow = pv["df"].iloc[mi]
                result_row += list(prow)
                if not matched:
                    matched = True
                    price_status, cdr_p, prov_p = compare_prices(
                        cdr_row, cdr_price_cols, prow, pv["price_cols"])
            else:
                result_row += [""] * len(pv["headers"])

            for pi in pv["price_cols"]:
                price_col_indices.add(offset + pi)
            offset += len(pv["headers"])

        for pi in price_col_indices:
            if pi < len(result_row) and result_row[pi] != "":
                result_row[pi] = fmt_money(result_row[pi])

        estado = {
            "mas_caro":   "CDR MÁS CARO",
            "mas_barato": "CDR MÁS BARATO",
            "igual":      "IGUAL",
            "sin_precio": "SIN PRECIO",
            "sin_match":  "SIN COINCIDENCIA",
        }.get(price_status, "")

        result_row += [
            estado,
            fmt_money(cdr_p)  if cdr_p  is not None else "",
            fmt_money(prov_p) if prov_p is not None else "",
        ]
        out_rows.append(result_row)
        row_meta.append({
            "matched":           matched,
            "price_status":      price_status,
            "price_col_indices": price_col_indices,
        })
        progress.progress((idx + 1) / total,
                          text=f"Procesando {idx+1}/{total}...")

    progress.empty()

    # Resumen
    con_match   = sum(1 for m in row_meta if m["matched"])
    sin_match   = total - con_match
    mas_caros   = sum(1 for m in row_meta if m["price_status"] == "mas_caro")
    mas_baratos = sum(1 for m in row_meta if m["price_status"] == "mas_barato")
    iguales     = sum(1 for m in row_meta if m["price_status"] == "igual")

    # Construir Excel en memoria
    cruce_df = pd.DataFrame(out_rows, columns=out_headers)
    resumen_df = pd.DataFrame({
        "Indicador": [
            "Total productos CDR", "Con coincidencia", "Sin coincidencia", "",
            "CDR más CARO", "CDR más BARATO", "Precio IGUAL",
        ],
        "Cantidad": [total, con_match, sin_match, "", mas_caros, mas_baratos, iguales],
        "Porcentaje / Descripción": [
            "100%",
            f"{con_match/total*100:.1f}%" if total else "0%",
            f"{sin_match/total*100:.1f}%" if total else "0%",
            "",
            "Precio CDR superior al proveedor",
            "Precio CDR inferior al proveedor",
            "Mismo precio en ambos",
        ],
    })

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        cruce_df.to_excel(writer,   sheet_name="Cruce CDR", index=False)
        resumen_df.to_excel(writer, sheet_name="Resumen",   index=False)
    buf.seek(0)

    # Aplicar estilos
    wb = load_workbook(buf)
    ws = wb["Cruce CDR"]
    for cell in ws[1]:
        cell.fill = FILL_HDR
        cell.font = FONT_HDR

    for row_idx, meta in enumerate(row_meta, start=2):
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            dc       = col_idx - 1
            is_price = dc in meta["price_col_indices"]
            if meta["matched"]:
                cell.fill = FILL_GREEN
            if is_price:
                if   meta["price_status"] == "mas_caro":   cell.fill = FILL_RED;    cell.font = FONT_RED
                elif meta["price_status"] == "mas_barato": cell.fill = FILL_BLUE;   cell.font = FONT_BLUE
                elif meta["price_status"] == "igual":      cell.fill = FILL_YELLOW; cell.font = FONT_YELLOW

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(len(str(c.value or "")) for c in col) + 4, 40)

    ws2 = wb["Resumen"]
    for cell in ws2[1]:
        cell.fill = FILL_HDR
        cell.font = FONT_HDR
    for col in ws2.columns:
        ws2.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(len(str(c.value or "")) for c in col) + 4, 50)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    return out, con_match, sin_match, mas_caros, mas_baratos, iguales, total


# ── UI Streamlit ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="Cruce CDR", page_icon="📊", layout="centered")
st.title("📊 Cruce CDR vs Proveedores")
st.markdown("Sube todos los archivos Excel. El que tenga **CDR** en el nombre será la base del cruce.")

uploaded = st.file_uploader(
    "Arrastra o selecciona archivos Excel",
    type=["xlsx", "xls"],
    accept_multiple_files=True,
)

if uploaded:
    cdr   = [f for f in uploaded if "CDR" in f.name.upper()]
    provs = [f for f in uploaded if "CDR" not in f.name.upper()]
    st.markdown(f"**{len(uploaded)} archivo(s) cargados** — "
                f"{'✅ CDR detectado' if cdr else '⚠️ Sin CDR'} | "
                f"{len(provs)} proveedor(es)")

    if st.button("🚀 Generar Cruce CDR", type="primary", disabled=not cdr or not provs):
        with st.spinner("Procesando..."):
            resultado = procesar(uploaded)

        if resultado:
            out, con_match, sin_match, mas_caros, mas_baratos, iguales, total = resultado

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total CDR",       total)
            col2.metric("Con coincidencia", con_match)
            col3.metric("Más caros",        mas_caros)
            col4.metric("Más baratos",      mas_baratos)

            st.success(f"✅ Listo — {con_match}/{total} productos con coincidencia")
            st.download_button(
                label="⬇️ Descargar Cruce_CDR.xlsx",
                data=out,
                file_name="Cruce_CDR.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
